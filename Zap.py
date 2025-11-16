from dotenv import load_dotenv
import os, sys, argparse, csv, json, datetime, re, time
from zoneinfo import ZoneInfo

try:
    import requests
except Exception as e:
    print("requests module is required.")
    sys.exit(1)

try:
    import openpyxl
except Exception:
    openpyxl = None

try:
    import xlsxwriter
except Exception:
    xlsxwriter = None

class ZendeskHttp:
    def __init__(self, sSub, sEmail, sTok):
        self.sBase = f"https://{sSub}.zendesk.com"
        self.o = requests.Session()
        self.o.auth = (f"{sEmail}/token", sTok)
        self.o.headers.update({"User-Agent": "Zap/1.0", "Accept": "application/json"})
        self.nTimeout = 30

    def _get_json(self, sUrl, nMaxRetries=6):
        nTry = 0
        while True:
            nTry += 1
            try:
                r = self.o.get(sUrl, timeout=self.nTimeout)
            except requests.RequestException as e:
                if nTry >= nMaxRetries:
                    print(f"Network error contacting Zendesk: {e}")
                    sys.exit(1)
                nSleep = min(2 ** (nTry - 1), 30)
                time.sleep(nSleep)
                continue

            if r.status_code == 429:
                sRetry = r.headers.get("Retry-After", "2")
                try:
                    nSleep = max(1, int(float(sRetry)))
                except Exception:
                    nSleep = 2
                if nTry >= nMaxRetries:
                    print("Rate limited by Zendesk too many times (429).")
                    sys.exit(1)
                time.sleep(nSleep)
                continue

            if 500 <= r.status_code < 600:
                if nTry >= nMaxRetries:
                    print(f"Zendesk server error {r.status_code}.")
                    try:
                        print(json.dumps(r.json(), ensure_ascii=False))
                    except Exception:
                        pass
                    sys.exit(1)
                nSleep = min(2 ** (nTry - 1), 30)
                time.sleep(nSleep)
                continue

            if r.status_code in (401, 403):
                try:
                    dErr = r.json()
                except Exception:
                    dErr = {}
                print(f"Authentication/authorization failed ({r.status_code}). Check ZENDESK_SUBDOMAIN / ZENDESK_EMAIL / ZENDESK_API_TOKEN.")
                if dErr:
                    print(json.dumps(dErr, ensure_ascii=False))
                sys.exit(1)

            try:
                r.raise_for_status()
            except requests.HTTPError as e:
                print(f"HTTP error from Zendesk: {e}")
                try:
                    print(json.dumps(r.json(), ensure_ascii=False))
                except Exception:
                    pass
                sys.exit(1)

            try:
                return r.json()
            except ValueError:
                print("Invalid JSON received from Zendesk.")
                sys.exit(1)

    def next_link(self, dJ):
        sL = None
        try:
            sL = dJ.get("links", {}).get("next")
        except Exception:
            sL = None
        if not sL:
            sL = dJ.get("next_page")
        return sL

class Resolver:
    def __init__(self, http: ZendeskHttp):
        self.h = http
        self.org = {}
        self.usr = {}
        self.grp = {}

    def _fetch_batch(self, kind, ids):
        ids = [str(i) for i in ids if i]
        if not ids:
            return
        sUrl = ""
        if kind == "org":
            sUrl = f"{self.h.sBase}/api/v2/organizations/show_many.json?ids={','.join(ids)}"
            d = self.h._get_json(sUrl)
            for x in d.get("organizations", []):
                self.org[str(x.get("id"))] = x.get("name") or ""
        elif kind == "usr":
            sUrl = f"{self.h.sBase}/api/v2/users/show_many.json?ids={','.join(ids)}"
            d = self.h._get_json(sUrl)
            for x in d.get("users", []):
                self.usr[str(x.get("id"))] = x.get("name") or ""
        elif kind == "grp":
            sUrl = f"{self.h.sBase}/api/v2/groups/show_many.json?ids={','.join(ids)}"
            d = self.h._get_json(sUrl)
            for x in d.get("groups", []):
                self.grp[str(x.get("id"))] = x.get("name") or ""

    def preload_from_tickets(self, aTickets):
        aOrg = set()
        aUsr = set()
        aGrp = set()
        for t in aTickets:
            if t.get("organization_id"): aOrg.add(str(t.get("organization_id")))
            if t.get("assignee_id"):     aUsr.add(str(t.get("assignee_id")))
            if t.get("requester_id"):    aUsr.add(str(t.get("requester_id")))
            if t.get("submitter_id"):    aUsr.add(str(t.get("submitter_id")))
            if t.get("group_id"):        aGrp.add(str(t.get("group_id")))
        self._fetch_batch("org", [i for i in aOrg if i not in self.org])
        self._fetch_batch("usr", [i for i in aUsr if i not in self.usr])
        self._fetch_batch("grp", [i for i in aGrp if i not in self.grp])

    def org_name(self, v): return self.org.get(str(v), "")
    def user_name(self, v): return self.usr.get(str(v), "")
    def group_name(self, v): return self.grp.get(str(v), "")

def buildWindowsUtc(sDateExpr, sStartTime, sEndTime):
    tzMnl = ZoneInfo("Asia/Manila")
    aDateRanges = []
    if sDateExpr:
        for part in re.split(r"\s+OR\s+", sDateExpr.strip(), flags=re.IGNORECASE):
            m = re.match(r"^\s*(\d{4}-\d{2}-\d{2})\s+TO\s+(\d{4}-\d{2}-\d{2})\s*$", part)
            if not m:
                print("Invalid date expression. Use: YYYY-MM-DD TO YYYY-MM-DD [OR YYYY-MM-DD TO YYYY-MM-DD]")
                sys.exit(1)
            aDateRanges.append((m.group(1), m.group(2)))

    aOut = []
    nowMnl = datetime.datetime.now(tzMnl)

    if not aDateRanges and not sStartTime and not sEndTime:
        t = nowMnl.time()
        def T(h, m): return datetime.time(hour=h, minute=m)
        if t < T(1,30):
            d = nowMnl.date()
            start = datetime.datetime.combine(d, T(13,0), tzMnl)
            end   = datetime.datetime.combine(d, T(1,30), tzMnl) + datetime.timedelta(days=1)
            aOut.append((start.astimezone(datetime.timezone.utc), end.astimezone(datetime.timezone.utc), "afternoon"))
        elif t < T(9,30):
            d = nowMnl.date()
            start = datetime.datetime.combine(d- datetime.timedelta(days=1), T(21,30), tzMnl)
            end   = datetime.datetime.combine(d, T(9,30), tzMnl)
            aOut.append((start.astimezone(datetime.timezone.utc), end.astimezone(datetime.timezone.utc), "evening"))
        elif t < T(18,30):
            d = nowMnl.date()
            start = datetime.datetime.combine(d, T(18,30), tzMnl)
            end   = datetime.datetime.combine(d+ datetime.timedelta(days=1), T(6,30), tzMnl)
            aOut.append((start.astimezone(datetime.timezone.utc), end.astimezone(datetime.timezone.utc), "morning"))
        else:
            d = nowMnl.date()
            start = datetime.datetime.combine(d, T(18,30), tzMnl)
            end   = datetime.datetime.combine(d+ datetime.timedelta(days=1), T(6,30), tzMnl)
            aOut.append((start.astimezone(datetime.timezone.utc), end.astimezone(datetime.timezone.utc), "morning"))
        return aOut

    if sStartTime or sEndTime:
        if not (sStartTime and sEndTime):
            print("Both --start and --end are required if time window is used.")
            sys.exit(1)
        try:
            tStart = datetime.datetime.strptime(sStartTime.strip(), "%I:%M %p").time()
            tEnd   = datetime.datetime.strptime(sEndTime.strip(), "%I:%M %p").time()
        except Exception:
            print("Invalid time. Use HH:MM AM/PM (e.g., 10:00 AM).")
            sys.exit(1)
        if (datetime.datetime.combine(datetime.date.today(), tEnd) <=
            datetime.datetime.combine(datetime.date.today(), tStart)):
            print("Time window must be within the day (end after start).")
            sys.exit(1)
    else:
        tStart = None
        tEnd = None

    if aDateRanges and sStartTime and sEndTime:
        if len(aDateRanges) != 1:
            print("If one date only and time range, supply a single date (YYYY-MM-DD TO YYYY-MM-DD with same start/end).")
            sys.exit(1)

    if not aDateRanges and (sStartTime and sEndTime):
        d = nowMnl.date()
        start = datetime.datetime.combine(d, tStart, tzMnl)
        end   = datetime.datetime.combine(d, tEnd, tzMnl)
        aOut.append((start.astimezone(datetime.timezone.utc), end.astimezone(datetime.timezone.utc), "time-only"))
        return aOut

    if aDateRanges and not (sStartTime or sEndTime):
        for d0, d1 in aDateRanges:
            sd = datetime.datetime.strptime(d0, "%Y-%m-%d").date()
            ed = datetime.datetime.strptime(d1, "%Y-%m-%d").date()
            start = datetime.datetime.combine(sd, datetime.time(0,0), tzMnl)
            end   = datetime.datetime.combine(ed, datetime.time(23,59,59), tzMnl)
            aOut.append((start.astimezone(datetime.timezone.utc), end.astimezone(datetime.timezone.utc), "date-only"))
        return aOut

    if aDateRanges and sStartTime and sEndTime and len(aDateRanges) == 1:
        d0, d1 = aDateRanges[0]
        if d0 != d1:
            print("If one date only and time range, the date range must be a single day.")
            sys.exit(1)
        dd = datetime.datetime.strptime(d0, "%Y-%m-%d").date()
        start = datetime.datetime.combine(dd, tStart, tzMnl)
        end   = datetime.datetime.combine(dd, tEnd, tzMnl)
        aOut.append((start.astimezone(datetime.timezone.utc), end.astimezone(datetime.timezone.utc), "date+time"))
        return aOut

    print("Invalid date/time combination.")
    sys.exit(1)

def harvestTicketsInWindows(oHttp, aWindowsUtc):
    aIds = []
    aAll = []
    seen = set()
    for (dtStartUtc, dtEndUtc, label) in aWindowsUtc:
        sStart = dtStartUtc.strftime("%Y-%m-%dT%H:%M:%SZ")
        sEnd   = dtEndUtc.strftime("%Y-%m-%dT%H:%M:%SZ")
        sQ = f"type:ticket created>={sStart} created<={sEnd}"
        sPage = f"{oHttp.sBase}/api/v2/search.json?query={requests.utils.quote(sQ, safe=':+=>=<')}&per_page=100"
        while sPage:
            dJ = oHttp._get_json(sPage)
            for r in dJ.get("results", []):
                if r.get("result_type") == "ticket":
                    tid = r.get("id")
                    if tid in seen:
                        continue
                    seen.add(tid)
                    aIds.append(str(tid))
            sPage = oHttp.next_link(dJ)

    if not aIds:
        return []

    nChunk = 100
    for i in range(0, len(aIds), nChunk):
        aChunkIds = aIds[i:i+nChunk]
        sUrl = f"{oHttp.sBase}/api/v2/tickets/show_many.json?ids={','.join(aChunkIds)}"
        dT = oHttp._get_json(sUrl)
        for t in dT.get("tickets", []):
            aAll.append(t)
    return aAll

def writeOutput(sBaseName, aRows, aHeaders, bMakeXlsx):
    sCsv = f"{sBaseName}.csv"
    with open(sCsv, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=aHeaders, extrasaction="ignore", quoting=csv.QUOTE_ALL, lineterminator="\r\n")
        w.writeheader()
        for row in aRows:
            dOut = {}
            for k in aHeaders:
                vRaw = row.get(k)
                if vRaw is None:
                    vCell = ""
                elif isinstance(vRaw, (dict, list)):
                    vCell = json.dumps(vRaw, ensure_ascii=False)
                elif isinstance(vRaw, str):
                    vCell = vRaw.replace("\r", " ").replace("\n", " ")
                else:
                    vCell = vRaw
                dOut[k] = vCell
            w.writerow(dOut)
    print(f"Wrote CSV -> {sCsv}")

    if bMakeXlsx:
        if xlsxwriter is None:
            print("xlsxwriter not installed, skipping workbook.")
            return
        sX = f"{sBaseName}.xlsx"
        wb = xlsxwriter.Workbook(sX, {"constant_memory": True})
        ws = wb.add_worksheet("tickets")
        fmtHead = wb.add_format({"bold": True, "border": 1, "text_wrap": True, "align": "center", "valign": "vcenter"})
        fmtVal  = wb.add_format({"border": 1, "text_wrap": True, "align": "left", "valign": "vcenter"})
        for j, h in enumerate(aHeaders):
            ws.write(0, j, h, fmtHead)
        for i, row in enumerate(aRows, start=1):
            for j, h in enumerate(aHeaders):
                vRaw = row.get(h)
                if vRaw is None:
                    vCell = ""
                elif isinstance(vRaw, (dict, list)):
                    vCell = json.dumps(vRaw, ensure_ascii=False)
                elif isinstance(vRaw, str):
                    vCell = vRaw.replace("\r", " ").replace("\n", " ")
                else:
                    vCell = vRaw
                ws.write(i, j, vCell, fmtVal)
        ws.set_row(0, 22)
        for j in range(len(aHeaders)):
            ws.set_column(j, j, 28)
        wb.close()
        print(f"Wrote XLSX -> {sX}")

def main():
    oP = argparse.ArgumentParser(
        prog="Zap.py",
        description="Export Zendesk tickets to CSV (and optional XLSX).",
        formatter_class=argparse.RawTextHelpFormatter
    )

    oP.add_argument("-f", "--fields", required=True, metavar="FILE",
                    help="Path to CSV or XLSX file listing Zendesk field names and IDs.")
    oP.add_argument("-c", "--creds", default=None, metavar="PATH",
                    help="Path to credentials .env (default: credentials.env in script directory).")
    oP.add_argument("--version", action="version", version="Zap 1.0")

    gFilt = oP.add_argument_group("Filters")
    gFilt.add_argument("-d", "--date", default=None, metavar="EXPR",
                       help="Date range expression using TO and OR.\nSample: 2025-01-01 TO 2025-01-10 OR 2025-02-01 TO 2025-02-05\nIf omitted and no time: default shift window based on current time in Manila.")
    gFilt.add_argument("-s", "--start", default=None, metavar="TIME",
                       help='Start time in HH:MM AM/PM (e.g., "10:00 AM").')
    gFilt.add_argument("-e", "--end", default=None, metavar="TIME",
                       help='End time in HH:MM AM/PM (e.g., "06:30 PM").')

    gOut = oP.add_argument_group("Output")
    gOut.add_argument("-o", "--output", default=None, metavar="NAME",
                      help="Output CSV filename (no extension). Default: autogenerated timestamp (yyyymmdd_hhmmss_am/pm).")
    gOut.add_argument("--xlsx", action="store_true",
                      help="Also export results to XLSX (same base filename as CSV).")

    a = oP.parse_args()

    sPath = a.fields
    aFlds = []
    if sPath.lower().endswith(".csv"):
        with open(sPath, "r", encoding="utf-8-sig", newline="") as f:
            oR = csv.reader(f)
            aRows = list(oR)
    else:
        if openpyxl is None:
            print("openpyxl required to read XLSX field list.")
            sys.exit(1)
        wb = openpyxl.load_workbook(sPath, read_only=True, data_only=True)
        ws = wb.active
        aRows = []
        for row in ws.iter_rows(values_only=True):
            aRows.append([(c if c is not None else "") for c in row])

    if not aRows:
        print("Empty fields file.")
        sys.exit(1)

    aHdr = [str(x).strip() for x in aRows[0]]

    try:
        iName = aHdr.index("Display name")
    except ValueError:
        try:
            iName = aHdr.index("Field")
        except ValueError:
            iName = 0

    try:
        iType = aHdr.index("Type")
    except ValueError:
        iType = None

    try:
        iId = aHdr.index("Field ID")
    except ValueError:
        try:
            iId = aHdr.index("ID")
        except ValueError:
            iId = None

    for r in aRows[1:]:
        sName = str(r[iName]).strip() if iName is not None and iName < len(r) else ""
        if not sName:
            continue
        sTyp  = str(r[iType]).strip() if iType is not None and iType < len(r) else ""
        sId   = str(r[iId]).strip() if iId is not None and iId < len(r) else ""
        aFlds.append((sName, sTyp, sId))

    print("Fields loaded:")
    for sName, sTyp, sId in aFlds:
        sT = f" [{sTyp}]" if sTyp else ""
        print(f"  {sName}{sT}")

    sCredsPath = a.creds
    if sCredsPath:
        if not os.path.isfile(sCredsPath):
            print("Credentials file not found.")
            sys.exit(1)
        load_dotenv(dotenv_path=sCredsPath, override=True)
    else:
        sScriptDir = os.path.dirname(os.path.abspath(__file__))
        sDefault = os.path.join(sScriptDir, "credentials.env")
        if not os.path.isfile(sDefault):
            print("Missing credentials.env in this folder. Create a file named credentials.env here with the following contents:")
            print("")
            print("ZENDESK_SUBDOMAIN=<Your Subdomain>")
            print("ZENDESK_EMAIL=<Your Email>")
            print("ZENDESK_API_TOKEN=<Your Token>")
            sys.exit(1)
        load_dotenv(dotenv_path=sDefault, override=True)
    sSub = os.getenv("ZENDESK_SUBDOMAIN")
    sEmail = os.getenv("ZENDESK_EMAIL")
    sTok = os.getenv("ZENDESK_API_TOKEN")
    if not all([sSub, sEmail, sTok]):
        print("Incomplete .env file...")
        print("")
        print("ZENDESK_SUBDOMAIN=<Your Subdomain>")
        print("ZENDESK_EMAIL=<Your Email>")
        print("ZENDESK_API_TOKEN=<Your Token>")
        sys.exit(1)

    http = ZendeskHttp(sSub, sEmail, sTok)

    dMe = http._get_json(f"{http.sBase}/api/v2/users/me.json")
    nMyId = dMe.get("user", {}).get("id")
    if not nMyId:
        print("Unexpected response from /users/me.json")
        print(json.dumps(dMe, ensure_ascii=False))
        sys.exit(1)

    aWindowsUtc = buildWindowsUtc(a.date, a.start, a.end)

    aTickets = harvestTicketsInWindows(http, aWindowsUtc)

    nChunkSize = 50
    aRowsOut = []
    dSeen = set()
    res = Resolver(http)

    aStdCols = ["ID", "Organization", "Assignee", "Group", "Status", "Type", "Subject", "Tags", "Created at", "Updated at"]
    aCustomCols = [s for (s, t, i) in aFlds]
    aHeaders = aStdCols + aCustomCols

    nWritten = 0
    for i in range(0, len(aTickets), nChunkSize):
        aChunk = aTickets[i:i+nChunkSize]
        res.preload_from_tickets(aChunk)

        for t in aChunk:
            tid = t.get("id")
            if tid in dSeen:
                continue
            dSeen.add(tid)

            dRow = {
                "ID": tid,
                "Organization": res.org_name(t.get("organization_id")),
                "Assignee": res.user_name(t.get("assignee_id")),
                "Group": res.group_name(t.get("group_id")),
                "Status": t.get("status") or "",
                "Type": t.get("type") or "",
                "Subject": t.get("subject") or "",
                "Tags": ",".join(t.get("tags") or []),
                "Created at": t.get("created_at") or "",
                "Updated at": t.get("updated_at") or "",
            }

            for sName, sTyp, sId in aFlds:
                v = ""
                if sId:
                    vCf = None
                    try:
                        for cf in t.get("custom_fields", []):
                            if str(cf.get("id")) == str(sId):
                                vCf = cf.get("value")
                                break
                    except Exception:
                        vCf = None
                    v = vCf if vCf is not None else ""
                dRow[sName] = v

            aRowsOut.append(dRow)
            nWritten += 1

        print(f"[{datetime.datetime.now()}] Chunk {(i//nChunkSize)+1}: scanned {len(aChunk)}, accumulated {nWritten}")

    if a.output:
        sBase = os.path.splitext(a.output)[0]
    else:
        nowMnl = datetime.datetime.now(ZoneInfo("Asia/Manila"))
        sBase = nowMnl.strftime("%Y%m%d_%I%M%S_%p").lower()

    writeOutput(sBase, aRowsOut, aHeaders, a.xlsx)

    if aTickets:
        sRaw = f"{sBase}.txt"
        with open(sRaw, "w", encoding="utf-8-sig") as f:
            for t in aTickets:
                f.write(json.dumps(t, ensure_ascii=False))
                f.write("\n")
        print(f"Wrote raw tickets -> {sRaw}")

if __name__ == "__main__":
    main()
